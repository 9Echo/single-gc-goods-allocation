import json
from app.main.dao.commodity_dao import write_db
from openpyxl import load_workbook, Workbook

def spiral_welded_pipe(itemid):
    """计算螺旋焊管的件重

    Args:
        tiemid:规格

    Returns:
        result: 件重

    Raise:

    """
    temp1 = itemid.split("*")
    outside = float((temp1[0][3:]).lstrip("0"))
    wall = float(temp1[1])
    length = float(temp1[2]) / 1000
    # (外径-(壁厚-0.9))*(壁厚-0.9)*0.0246615=米重(公斤)*12米=支重
    result = ((outside - (wall - 0.9)) * (wall - 0.9) * 0.0246615) * length
    return result


def rectangular_pipe(itemid):
    """计算方矩管的件重

    Args:
       itemid: 规格

    Returns:
        result: 件重

    Raise:

    """
    temp1 = itemid.split("*")
    width = float((temp1[0][3:]).lstrip("0"))
    height = float(temp1[1])
    wall = float(temp1[2])
    length = float(temp1[3]) / 1000
    # ((周长-3)/3.14159-壁厚)*壁厚*0.02466=米重*6米*整件支数=件重
    c = (width + height) * 2
    result = (((c - 3) / 3.14159 - wall) * wall * 0.02466) * length
    return result


def welded_pipe(itemid):
    """计算焊管的件重

    Args:
        itemid: 规格

    Returns:
        result: 件重

    Raise:

    """
    temp1 = itemid.split("*")
    outside = float((temp1[0][3:]).lstrip("0"))
    wall = float(temp1[1])
    length = float(temp1[2]) / 1000
    # (外径-壁厚)*壁厚*0.02466=米重*6米*整件支数=件重
    result = ((outside - wall) * wall * 0.02466) * length
    return result

def write_xlsx(data):
    wb = load_workbook("test1.xlsx")
    ws = wb.active
    row = ws.max_row
    ws.append(row+1, 1, data[0])
    ws.append(row+1, 2, data[1])
    ws.append(row+1, 3, data[2])
    ws.append(row+1, 4, data[3])


with open("json.txt", "rt") as f:
    temp = json.loads(f.read())
f.close()

for i in temp["data"]["list"]:
    cname = i["cname"]
    itemid = i["itemid"]
    endqty = float(i["endqty"])
    zg00 = float(i["zg00"])
    if zg00 != 0:
        div_result = endqty / zg00
    else:
        div_result = "总根数为0"
    list1 = itemid.split("*")
    if cname == "焊管":
        result = welded_pipe(itemid)
        write_xlsx([cname, itemid, div_result, result])
    if cname == "螺旋焊管":
        result = spiral_welded_pipe(itemid)
        write_xlsx([cname, itemid, div_result, result])
    if cname == "方矩管" or cname == "热镀方矩管":
        result = rectangular_pipe(itemid)
        write_xlsx([cname, itemid, div_result, result])










